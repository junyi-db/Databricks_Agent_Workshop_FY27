"""
Flatten the multi-sheet assets.xlsx into a single CSV with a business_unit
column. This is used to seed the validation workspace via SQL read_files,
while the original XLSX stays in the Volume for the workshop's hands-on
XLSX ingestion lab.
"""
from pathlib import Path
import csv
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "xlsx" / "assets.xlsx"
DST = ROOT / "data" / "csv" / "assets_flat.csv"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
header = None
out_rows = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    if header is None:
        header = list(rows[0]) + ["business_unit_sheet"]
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out_rows.append(list(r) + [sheet_name])

with DST.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(header)
    w.writerows(out_rows)

print(f"Wrote {DST.name}: {len(out_rows)} rows")
