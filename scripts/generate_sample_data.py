"""
Generate synthetic PETRONAS Operations Intelligence sample data for the workshop.

Outputs:
  data/csv/operations_daily.csv      - Daily operational metrics across BUs
  data/csv/commercial_summary.csv    - Cargo/sales/trade volumes
  data/xlsx/assets.xlsx              - Asset register, one sheet per BU

All data is synthetic and PETRONAS-shaped but does not represent real operations.
"""

from pathlib import Path
import random
import datetime as dt
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)
ROOT = Path(__file__).resolve().parents[1]
CSV_DIR = ROOT / "data" / "csv"
XLSX_DIR = ROOT / "data" / "xlsx"
CSV_DIR.mkdir(parents=True, exist_ok=True)
XLSX_DIR.mkdir(parents=True, exist_ok=True)

# -------------------------------------------------------------------
# Site master — covers all major PETRONAS BUs
# -------------------------------------------------------------------
SITES = [
    # Upstream (offshore platforms + onshore fields)
    {"site_id": "UP-BKSP-A", "site_name": "Bokor Platform A",          "business_unit": "Upstream",  "site_type": "Offshore Platform",   "country": "Malaysia",  "lat": 4.78,  "lng": 113.10, "throughput_unit": "kbbl/d", "target": 42.0},
    {"site_id": "UP-KSPL-1", "site_name": "Kasawari LNG Platform",     "business_unit": "Upstream",  "site_type": "Offshore Platform",   "country": "Malaysia",  "lat": 5.92,  "lng": 111.74, "throughput_unit": "MMscf/d", "target": 900.0},
    {"site_id": "UP-PMOC-3", "site_name": "PMO Cluster 3",             "business_unit": "Upstream",  "site_type": "Offshore Platform",   "country": "Malaysia",  "lat": 5.45,  "lng": 104.20, "throughput_unit": "kbbl/d", "target": 28.0},
    {"site_id": "UP-CYNB-1", "site_name": "Carigali Onshore Block",    "business_unit": "Upstream",  "site_type": "Onshore Field",       "country": "Malaysia",  "lat": 4.20,  "lng": 113.99, "throughput_unit": "kbbl/d", "target": 12.0},
    # Downstream (refineries + petrochem)
    {"site_id": "DS-MLKR-1", "site_name": "Melaka Refinery PSR-1",     "business_unit": "Downstream","site_type": "Refinery",            "country": "Malaysia",  "lat": 2.36,  "lng": 102.16, "throughput_unit": "kbbl/d", "target": 100.0},
    {"site_id": "DS-MLKR-2", "site_name": "Melaka Refinery PSR-2",     "business_unit": "Downstream","site_type": "Refinery",            "country": "Malaysia",  "lat": 2.36,  "lng": 102.16, "throughput_unit": "kbbl/d", "target": 170.0},
    {"site_id": "DS-PNGRN-1","site_name": "Pengerang RAPID Refinery",  "business_unit": "Downstream","site_type": "Refinery",            "country": "Malaysia",  "lat": 1.39,  "lng": 104.13, "throughput_unit": "kbbl/d", "target": 300.0},
    {"site_id": "DS-KRTH-1", "site_name": "Kerteh Petrochemical",      "business_unit": "Downstream","site_type": "Petrochemical Plant", "country": "Malaysia",  "lat": 4.51,  "lng": 103.45, "throughput_unit": "kt/d",   "target": 8.0},
    # Gas & Power (LNG)
    {"site_id": "GP-BTLG-1", "site_name": "Bintulu LNG Complex T1",    "business_unit": "Gas",       "site_type": "LNG Plant",           "country": "Malaysia",  "lat": 3.27,  "lng": 113.07, "throughput_unit": "MTPA",   "target": 3.4},
    {"site_id": "GP-BTLG-2", "site_name": "Bintulu LNG Complex T9",    "business_unit": "Gas",       "site_type": "LNG Plant",           "country": "Malaysia",  "lat": 3.27,  "lng": 113.07, "throughput_unit": "MTPA",   "target": 3.6},
    # Trading
    {"site_id": "TR-KL-DSK", "site_name": "PETCO Trading Desk KL",     "business_unit": "Trading",   "site_type": "Trading Desk",        "country": "Malaysia",  "lat": 3.16,  "lng": 101.71, "throughput_unit": "kbbl/d", "target": 250.0},
    {"site_id": "TR-LDN-DSK","site_name": "PETCO Trading Desk London", "business_unit": "Trading",   "site_type": "Trading Desk",        "country": "UK",        "lat": 51.51, "lng": -0.13,  "throughput_unit": "kbbl/d", "target": 180.0},
    # Retail (fuel stations - aggregated cluster levels)
    {"site_id": "RT-KLG-CL", "site_name": "Klang Valley Cluster",      "business_unit": "Retail",    "site_type": "Retail Cluster",      "country": "Malaysia",  "lat": 3.13,  "lng": 101.62, "throughput_unit": "kL/d",   "target": 950.0},
    {"site_id": "RT-PNG-CL", "site_name": "Penang Cluster",            "business_unit": "Retail",    "site_type": "Retail Cluster",      "country": "Malaysia",  "lat": 5.41,  "lng": 100.33, "throughput_unit": "kL/d",   "target": 480.0},
    {"site_id": "RT-JHR-CL", "site_name": "Johor Cluster",             "business_unit": "Retail",    "site_type": "Retail Cluster",      "country": "Malaysia",  "lat": 1.49,  "lng": 103.74, "throughput_unit": "kL/d",   "target": 410.0},
    # Logistics
    {"site_id": "LG-PSGD-T", "site_name": "Pasir Gudang Marine Terminal","business_unit": "Logistics","site_type": "Marine Terminal",    "country": "Malaysia",  "lat": 1.46,  "lng": 103.89, "throughput_unit": "kbbl/d", "target": 220.0},
    {"site_id": "LG-LBN-T",  "site_name": "Labuan Crude Terminal",     "business_unit": "Logistics", "site_type": "Marine Terminal",     "country": "Malaysia",  "lat": 5.28,  "lng": 115.24, "throughput_unit": "kbbl/d", "target": 140.0},
]

# -------------------------------------------------------------------
# operations_daily.csv — last 6 months of daily metrics per site
# -------------------------------------------------------------------
def gen_operations_daily():
    end_date = dt.date(2026, 4, 30)
    start_date = end_date - dt.timedelta(days=180)
    rows = []
    rid = 1
    for site in SITES:
        # Per-site emission factor (tCO2e per unit throughput)
        emission_factor = {
            "Offshore Platform": 0.45,
            "Onshore Field": 0.30,
            "Refinery": 0.55,
            "Petrochemical Plant": 0.85,
            "LNG Plant": 0.40,
            "Trading Desk": 0.05,
            "Retail Cluster": 0.10,
            "Marine Terminal": 0.15,
        }[site["site_type"]]

        # Each site has a slightly different operational profile
        base_uptime = random.uniform(0.92, 0.99)
        d = start_date
        while d <= end_date:
            # Realistic variability + a few outage events
            outage = 1 if random.random() < 0.015 else 0
            uptime = max(0.0, min(1.0, random.gauss(base_uptime, 0.025))) if not outage else random.uniform(0.0, 0.6)
            # Throughput correlated with uptime + small daily noise
            throughput = round(site["target"] * uptime * random.uniform(0.92, 1.08), 2)
            energy_mwh = round(throughput * random.uniform(2.5, 4.5), 1)
            emissions_tco2e = round(throughput * emission_factor * random.uniform(0.9, 1.1), 2)
            rows.append({
                "record_id": rid,
                "operation_date": d.isoformat(),
                "site_id": site["site_id"],
                "site_name": site["site_name"],
                "business_unit": site["business_unit"],
                "site_type": site["site_type"],
                "country": site["country"],
                "throughput_value": throughput,
                "throughput_unit": site["throughput_unit"],
                "throughput_target": site["target"],
                "uptime_pct": round(uptime * 100, 2),
                "energy_consumed_mwh": energy_mwh,
                "emissions_tco2e": emissions_tco2e,
                "outage_flag": outage,
                "latitude": site["lat"],
                "longitude": site["lng"],
            })
            rid += 1
            d += dt.timedelta(days=1)

    out = CSV_DIR / "operations_daily.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"  Wrote {out.name}: {len(rows):,} rows")


# -------------------------------------------------------------------
# commercial_summary.csv — monthly trade/cargo/retail commercial volumes
# -------------------------------------------------------------------
def gen_commercial():
    months = []
    d = dt.date(2025, 11, 1)
    while d <= dt.date(2026, 4, 1):
        months.append(d)
        # advance one month
        if d.month == 12:
            d = dt.date(d.year + 1, 1, 1)
        else:
            d = dt.date(d.year, d.month + 1, 1)

    products = ["Crude Oil", "Naphtha", "Gasoline", "Diesel", "Jet A-1", "LNG", "LPG", "Lubricants"]
    rows = []
    rid = 1
    for site in SITES:
        if site["business_unit"] not in {"Trading", "Retail", "Downstream", "Logistics"}:
            continue
        for m in months:
            for p in random.sample(products, k=random.randint(2, 4)):
                volume_kbbl = round(random.uniform(50, 1500), 1)
                price_usd = round(random.uniform(60, 95), 2)
                rev_musd = round(volume_kbbl * 1000 * price_usd / 1e6, 2)
                rows.append({
                    "record_id": rid,
                    "month": m.isoformat(),
                    "site_id": site["site_id"],
                    "business_unit": site["business_unit"],
                    "product": p,
                    "volume_kbbl": volume_kbbl,
                    "avg_price_usd_per_bbl": price_usd,
                    "revenue_musd": rev_musd,
                })
                rid += 1
    out = CSV_DIR / "commercial_summary.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"  Wrote {out.name}: {len(rows):,} rows")


# -------------------------------------------------------------------
# assets.xlsx — multi-sheet asset register (one sheet per BU)
# -------------------------------------------------------------------
def gen_assets_xlsx():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet

    asset_types_by_bu = {
        "Upstream":   ["Wellhead", "Topside Module", "Subsea Tree", "Riser", "Compressor", "Separator"],
        "Downstream": ["CDU", "VDU", "Reformer", "FCC Unit", "Hydrocracker", "Heat Exchanger"],
        "Gas":        ["LNG Train", "Liquefaction Unit", "BOG Compressor", "LNG Storage Tank", "Loading Arm"],
        "Trading":    ["Trading Workstation", "Risk Engine", "Storage Lease", "Vessel Charter Slot"],
        "Retail":     ["Forecourt Pump", "Underground Tank", "POS Terminal", "Convenience Store Unit"],
        "Logistics":  ["Berth", "Tank Farm", "Pipeline Segment", "Vapour Recovery Unit"],
    }

    asset_id_counter = 1000
    for bu, asset_types in asset_types_by_bu.items():
        bu_sites = [s for s in SITES if s["business_unit"] == bu]
        if not bu_sites:
            continue
        ws = wb.create_sheet(bu)
        headers = [
            "asset_id", "asset_name", "asset_type", "site_id", "site_name",
            "criticality", "install_year", "last_inspection_date",
            "next_inspection_date", "status", "owner_team",
        ]
        ws.append(headers)
        # Style header row
        for col_idx, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="00A19A")  # PETRONAS teal-ish
            c.alignment = Alignment(horizontal="center")

        n_assets = random.randint(15, 35)
        for _ in range(n_assets):
            site = random.choice(bu_sites)
            atype = random.choice(asset_types)
            asset_id_counter += 1
            install_year = random.randint(1995, 2024)
            last_insp = dt.date(2026, 4, 30) - dt.timedelta(days=random.randint(15, 720))
            interval_days = random.choice([180, 365, 730])
            next_insp = last_insp + dt.timedelta(days=interval_days)
            status = random.choices(
                ["In Service", "In Service", "In Service", "Under Maintenance", "Decommissioned"],
                weights=[60, 15, 15, 8, 2],
            )[0]
            ws.append([
                f"AST-{bu[:2].upper()}-{asset_id_counter}",
                f"{atype} {asset_id_counter}",
                atype,
                site["site_id"],
                site["site_name"],
                random.choice(["Critical", "High", "Medium", "Low"]),
                install_year,
                last_insp.isoformat(),
                next_insp.isoformat(),
                status,
                f"{bu} Asset Integrity",
            ])
        # Auto-width-ish
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 22 if col_idx in (2, 5) else 16

    out = XLSX_DIR / "assets.xlsx"
    wb.save(out)
    print(f"  Wrote {out.name}: {len(wb.sheetnames)} sheets ({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    print("Generating PETRONAS Ops Intelligence sample data...")
    gen_operations_daily()
    gen_commercial()
    gen_assets_xlsx()
    print("Done.")
